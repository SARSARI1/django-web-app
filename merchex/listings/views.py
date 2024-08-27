# ~/projects/django-web-app/merchex/listings/views.py
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .forms import AffectationForm, UploadFilesForm, UploadFilesFormLibre
from .models import DemandesTraiter, Quota
from django.core.files.storage import default_storage
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from io import BytesIO
import pandas as pd
from datetime import datetime
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook

from dateutil.relativedelta import relativedelta
from reportlab.lib.pagesizes import A4


# Sample view for hello.html
def hello(request):
    bands = Band.objects.all()
    return render(request, 'listings/hello.html', {'bands': bands})

# View for displaying tables.html
def tables(request):
    # Your logic for this view
    return render(request, 'listings/tables.html')

def dashboard(request):
    return render(request, 'listings/dashboard.html')

def demandes(request):
    return render(request, 'listings/demandes.html')
    
def quota(request):
    quotas = Quota.objects.all()
    return render(request, 'listings/quota.html',{'quotas': quotas})

def agents(request):
    return render(request, 'listings/agents.html')

def historique(request):
    historiques=Historique.objects.all()
    return render(request, 'listings/historique.html',{'historiques': historiques})



# View for about page
def about(request):
    return HttpResponse('<h1>À propos</h1> <p>Nous adorons merch !</p>')

# View for uploading files
from django.shortcuts import redirect, render
from django.contrib import messages
import pandas as pd
from .models import Agent, Demande, Historique
from .forms import UploadFilesForm,CalculForm
import pandas as pd
from django.shortcuts import redirect, render
from django.contrib import messages
from .models import Agent, Demande, Historique
from .forms import UploadFilesForm

def upload_files(request):
    if request.method == 'POST':
        form = UploadFilesForm(request.POST, request.FILES)
        if form.is_valid():
            # Extract files from the form
            agents_file = request.FILES.get('agents_file')
            historique_file = request.FILES.get('historique_file')
            demandes_file = request.FILES.get('demandes_file')

            # Process the agents file
            if agents_file:
                if not agents_file.name.endswith('.xlsx'):
                    messages.error(request, "Format de fichier invalide pour les agents. Veuillez télécharger un fichier .xlsx.")
                else:
                    try:
                        df = pd.read_excel(agents_file)
                        expected_columns = ['matricule', 'nom_prenom', 'date_naissance', 'sit_fam', 'date_embauche', 'nombre_enf', 'date_debut_retraite']
                        df.columns = df.columns.str.lower()
                        if not all(col in df.columns for col in expected_columns):
                            messages.error(request, "Format de fichier Excel invalide pour les agents. Les colonnes attendues sont manquantes.")
                        else:
                            # Delete all existing records from the Agent table
                            Agent.objects.all().delete()

                            for index, row in df.iterrows():
                                if pd.isnull(row['matricule']) or pd.isnull(row['nom_prenom']) or pd.isnull(row['date_naissance']):
                                    messages.error(request, f"Données requises manquantes à la ligne {index + 1}.")
                                    return redirect('AgentsAffichage')
                                try:
                                    agent = Agent(
                                        matricule=row['matricule'],
                                        nom_prenom=row['nom_prenom'],
                                        date_naissance=row['date_naissance'] if not pd.isna(row['date_naissance']) else None,
                                        sit_fam=row['sit_fam'],
                                        date_embauche=row['date_embauche'] if not pd.isna(row['date_embauche']) else None,
                                        nombre_enf=row['nombre_enf'],
                                        date_debut_retraite=row.get('date_debut_retraite') if not pd.isna(row.get('date_debut_retraite')) else None
                                    )
                                    agent.save()
                                except Exception as e:
                                    messages.error(request, f"Erreur lors de l'enregistrement des données des agents : {str(e)}")
                                    return redirect('AgentsAffichage')
                            messages.success(request, "Fichier agents téléchargé et traité avec succès !")
                    except Exception as e:
                        messages.error(request, f"Erreur lors de la lecture du fichier Excel des agents : {str(e)}")
                        return redirect('AgentsAffichage')

            # Process the demandes file
            if demandes_file:
                try:
                    df = pd.read_excel(demandes_file)
                    required_columns = ['Ville', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites', 'Statut', 'Nature Periode', 'Site', 'N° Demande']
                    for column in required_columns:
                        if column not in df.columns:
                            messages.error(request, f'Colonne manquante dans le fichier Excel des demandes: {column}')
                            return redirect('demande_list')

                    # Delete all existing records from the Demande table
                    Demande.objects.all().delete()

                    for _, row in df.iterrows():
                        Demande.objects.create(
                            ville=row.get('Ville', ''),
                            nom_agent=row.get('Nom agent', ''),
                            prenom_agent=row.get('Prenom agent', ''),
                            matricule=row.get('Matricule', ''),
                            date_demande=row.get('Date de la demande', pd.NaT),
                            date_debut_sejour=row.get('Date debut sejour', pd.NaT),
                            date_fin_sejour=row.get('Date fin sejour', pd.NaT),
                            type_de_vue=row.get('Type de vue', 'Inconnu'),
                            nombre_nuites=row.get('Nombre de nuites', 0),
                            statut=row.get('Statut', ''),
                            nature_periode=row.get('Nature Periode', ''),
                            site=row.get('Site', ''),
                            numero_demande=row.get('N° Demande', '')
                        )
                    messages.success(request, 'Fichier demandes téléchargé et données insérées avec succès.')
                except Exception as e:
                    messages.error(request, f'Erreur lors du traitement du fichier Excel des demandes: {e}')
                    return redirect('demande_list')

            # Process the historique file
            if historique_file:
                if not historique_file.name.endswith('.xlsx'):
                    messages.error(request, "Format de fichier invalide pour l'historique. Veuillez télécharger un fichier .xlsx.")
                else:
                    try:
                        df = pd.read_excel(historique_file)
                        required_columns = ['Site', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites']
                        if not all(column in df.columns for column in required_columns):
                            messages.error(request, "Le fichier Excel de l'historique n'a pas les colonnes requises.")
                            return redirect('historique')

                        # Delete all existing records from the Historique table
                        Historique.objects.all().delete()

                        rows_with_missing_values = df[df[required_columns].isnull().any(axis=1)]
                        df = df.dropna(subset=required_columns)
                        inserted_rows = 0
                        for _, row in df.iterrows():
                            try:
                                Historique.objects.create(
                                    ville=row['Site'],
                                    nom_agent=row['Nom agent'],
                                    prenom_agent=row['Prenom agent'],
                                    matricule=row['Matricule'],
                                    date_demande=row['Date de la demande'],
                                    date_debut_sejour=row['Date debut sejour'],
                                    date_fin_sejour=row['Date fin sejour'],
                                    type_de_vue=row['Type de vue'],
                                    nombre_nuites=row['Nombre de nuites'],
                                )
                                inserted_rows += 1
                            except Exception as e:
                                messages.error(request, f"Erreur lors de l'insertion de la ligne {row}: {e}")
                        if not rows_with_missing_values.empty:
                            messages.warning(request, "Certaines lignes avaient des valeurs manquantes et n'ont pas été insérées.", extra_tags='missing-values')
                        else:
                            messages.success(request, f"Fichier Excel historique traité avec succès. {inserted_rows} lignes insérées.")
                    except Exception as e:
                        messages.error(request, f"Erreur lors de la lecture du fichier Excel de l'historique : {str(e)}")
                        return redirect('historique')

            return render(request, 'listings/tables.html')
    else:
        return render(request, 'listings/tables.html')



# View for listing generated data
from django.utils.dateparse import parse_date




def list_generated(request):
    # Retrieve missing data from session
    missing_data = request.session.get('missing_data', None)

    # Fetch distinct combinations of 'ville' and 'type_de_vue'
    data = DemandesTraiter.objects.values('ville', 'type_de_vue').distinct()

    # Group data by 'ville' and store unique 'type_de_vue' values
    grouped_data = defaultdict(set)
    for item in data:
        grouped_data[item['ville']].add(item['type_de_vue'])

    # Convert set to list for rendering
    context = {
        'data': {ville: list(types) for ville, types in grouped_data.items()}
    }

    # If missing data exists, add it to the context and optionally clear the session
    if missing_data:
        context['missing_data'] = missing_data
        del request.session['missing_data']  # Clear the session variable after use

    return render(request, 'listings/tables.html', context)


# View for downloading PDF
def download_pdf(request, ville, type_de_vue):
    # Retrieve data from the database
    data = get_data_from_database(ville, type_de_vue)
    df = pd.DataFrame(data)
    
    # Print column names to debug
    print("DataFrame columns:", df.columns.tolist())
    
    # Sort the DataFrame based on 'P' in descending order
    sorted_table = df.sort_values(by=['P'], ascending=[False]).reset_index(drop=True)
    
    # Define columns of interest based on the model
    columns_of_interest = ['numero_demande', 'ville', 'nom_agent', 'prenom_agent', 'matricule',
                           'date_debut_sejour', 'date_fin_sejour', 'type_de_vue', 'P']
    columns_of_interest_print = ['numero_demande', 'nom_agent', 'prenom_agent', 'matricule',
                                 'date_debut_sejour', 'date_fin_sejour','A','S','D', 'P']

    # Check if all columns exist in the DataFrame
    missing_cols = [col for col in columns_of_interest if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns: {missing_cols}")

    # Prepare the selected data
    selected_data = sorted_table[columns_of_interest_print].copy()
    selected_data['date_debut_sejour'] = pd.to_datetime(selected_data['date_debut_sejour']).dt.strftime('%Y-%m-%d')
    selected_data['date_fin_sejour'] = pd.to_datetime(selected_data['date_fin_sejour']).dt.strftime('%Y-%m-%d')

    # Get the quota value
    quota_value = get_quota_value(ville, type_de_vue)

    # Separate data into principal and waiting lists based on quota
    if (quota_value or 0) >= len(selected_data):
        principal_data = selected_data
        waiting_data = pd.DataFrame(columns=columns_of_interest_print)
    else:
        principal_data = selected_data.head(quota_value)
        waiting_data = selected_data.tail(len(selected_data) - quota_value)

    # Convert DataFrames to lists for PDF generation
    principal_data_list = [principal_data.columns.tolist()] + principal_data.values.tolist()
    waiting_data_list = [waiting_data.columns.tolist()] + waiting_data.values.tolist()

    # Define titles for the PDF
    title_text = f"Liste principale pour {ville} ({type_de_vue})"
    title_text_wait = f"Liste d'attente pour {ville} ({type_de_vue})"

    # Create the PDF
    pdf_data = create_pdf(principal_data_list, waiting_data_list, calculate_column_widths(selected_data), title_text, title_text_wait)
    
    # Return the PDF response
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{ville}_{type_de_vue}.pdf"'
    
    return response


# Helper function to retrieve data from database
def get_data_from_database(ville, type_de_vue):
    return DemandesTraiter.objects.filter(ville=ville, type_de_vue=type_de_vue).values()


# Helper function to retrieve quota value
def get_quota_value(ville, type_de_vue):
    try:
        quota = Quota.objects.get(ville=ville, type_de_vue=type_de_vue)
        return quota.quota_value
    except Quota.DoesNotExist:
        return 0





# Helper function to calculate column widths for the table
def calculate_column_widths(data):
    col_widths = []
    for col in data.columns:
        if col != 'Type de vue':  # Exclude 'Type de vue' column
            max_len = max(data[col].astype(str).map(len).max(), len(col))  # Include header length
            col_widths.append(max_len * 8)  # Multiply by a factor to adjust width
    return col_widths



# Function to process uploaded files
from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd

# Define a function to convert DataFrame rows to Django model instances and save to the database


def save_to_database(df, request):
    try:
        # Delete all existing records in the DemandesTraiter table
        DemandesTraiter.objects.all().delete()
        
        demandes_traiter_list = []
        for _, row in df.iterrows():
            demande = DemandesTraiter(
                numero_demande=row['N° Demande'],
                ville=row['Ville'],
                nom_agent=row['Nom agent'],
                prenom_agent=row['Prenom agent'],
                matricule=row['Matricule'],
                date_debut_sejour=row['Date debut sejour'],
                date_fin_sejour=row['Date fin sejour'],
                type_de_vue=row['Type de vue'],
                A=row['A'],
                D=row['D'],
                S=row['S'],
                P=row['P'],
            )
            demandes_traiter_list.append(demande)

        # Bulk create all instances at once
        DemandesTraiter.objects.bulk_create(demandes_traiter_list)

        # Add a success message
        messages.success(request, "Les données ont été sauvegardées avec succès (Consultez les tableaux du tableau de bord).")
    except Exception as e:
        # Add an error message
        messages.error(request, f"Une erreur est survenue lors de l'enregistrement des données : {e}")



from django.utils.dateparse import parse_date

from django.contrib import messages
import pandas as pd
from listings.models import RejectedDemandesRetrait
def save_rejected_records(df, request):
    # Ensure date columns are in datetime format
    df['date_debut_retraite'] = pd.to_datetime(df['date_debut_retraite'], errors='coerce')
    df['Date debut sejour'] = pd.to_datetime(df['Date debut sejour'], errors='coerce')

    # Extract rejected records
    rejected_df = df[
        (df['date_debut_retraite'] <= df['Date debut sejour']) &
        (~df['date_debut_retraite'].isna())
    ]

    # Message about the number of rejected records
    messages.info(request, f"Nombre de demandes rejetées : {len(rejected_df)}")

    # Prepare a list for bulk creation
    rejected_records = []
    for _, row in rejected_df.iterrows():
        # Handle potential NaN values
        rejected_record = RejectedDemandesRetrait(
            numero_demande=row.get('N° Demande', ''),
            ville=row.get('Ville', ''),
            nom_agent=row.get('Nom agent', ''),
            prenom_agent=row.get('Prenom agent', ''),
            matricule=row.get('Matricule', ''),
            date_debut_sejour=row.get('Date debut sejour', pd.NaT),
            date_fin_sejour=row.get('Date fin sejour', pd.NaT),
            type_de_vue=row.get('Type de vue', ''),
            date_debut_retraite=row.get('date_debut_retraite', pd.NaT),
            date_de_la_demande=row.get('Date de la demande', pd.NaT)
        )
        rejected_records.append(rejected_record)

    # Delete all existing records in the table
    RejectedDemandesRetrait.objects.all().delete()

    # Bulk create all instances at once
    if rejected_records:
        RejectedDemandesRetrait.objects.bulk_create(rejected_records)
        messages.success(request, f"{len(rejected_records)} enregistrements ont été ajoutés avec succès (Consultez les tableaux du tableau de bord).")
    else:
        messages.info(request, "Aucun enregistrement à insérer.")




def proceed_with_calculation(request):
    # Retrieve dates from the session
    date_debut_sejour_str = request.session.get('date_debut_sejour')
    date_fin_sejour_str = request.session.get('date_fin_sejour')

    if date_debut_sejour_str and date_fin_sejour_str:
        # Convert strings back to date objects
        date_debut_sejour = parse_date(date_debut_sejour_str)
        date_fin_sejour = parse_date(date_fin_sejour_str)

        print("i got to inside")
       

        # Check if dates are provided
        if not date_debut_sejour or not date_fin_sejour:
            messages.error(request, "Les dates de début et de fin de séjour sont requises.")
            return redirect('list_generated')  # Replace with the actual view name
        
        try:
            # Fetch data from the models
            historique_queryset = Historique.objects.all().values()
            agents_queryset = Agent.objects.all().values()
            demandes_queryset = Demande.objects.all().values()
            # Convert the querysets to DataFrames
            historique_df = pd.DataFrame(list(historique_queryset))
            agents_df = pd.DataFrame(list(agents_queryset))
            demandes_df = pd.DataFrame(list(demandes_queryset))

            # Column name mappings
            historique_rename_dict = {
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'nombre_nuites': 'Nombre de nuites',
            }
            historique_df.rename(columns=historique_rename_dict, inplace=True)

            agents_rename_dict = {
                'matricule': 'matricule',
                'nom_prenom': 'nom_prenom',
                'date_naissance': 'date_naissance',
                'date_embauche': 'date_embauche',
                'nombre_enf': 'NOMBRE_ENF',
                'date_debut_retraite': 'date_debut_retraite',
            }
            agents_df.rename(columns=agents_rename_dict, inplace=True)

            demandes_rename_dict = {
                'numero_demande': 'N° Demande',
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'type_de_vue': 'Type de vue',
                'nombre_nuites': 'Nombre de nuites',
                'statut': 'Statut',
                'nature_periode': 'Nature Periode',
                'site': 'Site'
            }
            demandes_df.rename(columns=demandes_rename_dict, inplace=True)

            # Filter rows based on date_debut_sejour and date_fin_sejour
            demandes_df = demandes_df[
                (demandes_df['Date debut sejour'] == date_debut_sejour) &
                (demandes_df['Date fin sejour'] == date_fin_sejour)
            ]

            # Create a new DataFrame 'demandes_traiter' that contains the exact content of the 'demandes' file
            demandes_traiter_df = demandes_df.copy()

            # Step 4: Filter Rows
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Bloquée']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

            # Step 5: Add retraite Column and Calculate Its Values
            from datetime import datetime
            demandes_traiter_df = demandes_traiter_df.merge(
                agents_df[['matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF', 'date_debut_retraite']],
                left_on='Matricule', right_on='matricule', how='left'
            )

            # Convert dates to datetime
            demandes_traiter_df['date_embauche'] = pd.to_datetime(demandes_traiter_df['date_embauche'], format='%d/%m/%Y')
            demandes_traiter_df['Date debut sejour'] = pd.to_datetime(demandes_traiter_df['Date debut sejour'], format='%d/%m/%Y')
            demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')
            
            # Check for missing values in required columns of demandes_traiter_df
            required_columns = ['Matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF']
            
            demandes_traiter_df = demandes_traiter_df.dropna(subset=required_columns)
            messages.success(request, "L'opération a été effectuée avec succès après avoir ignoré les données manquantes. Les calculs ont été poursuivis avec les données disponibles.")
                

            # Filter expired date_debut_retraite
            save_rejected_records(demandes_traiter_df, request)
            demandes_traiter_df = demandes_traiter_df[
                (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                (demandes_traiter_df['date_debut_retraite'].isna())
            ]

            if demandes_traiter_df.empty:
                messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                return redirect('list_generated')

          

            # Define the fixed date: June 1st of the current year
            fixed_date = datetime(datetime.now().year, 6, 1)

            # Calculate the difference in months
            demandes_traiter_df['A'] = demandes_traiter_df['date_embauche'].apply(
                lambda x: relativedelta(fixed_date, x).years * 12 + relativedelta(fixed_date, x).months
            )


            # Calculate S
            def calculate_S(row):
                sit_fam = row['sit_fam'].strip().lower()
                nbr_enf = row['NOMBRE_ENF']

                if sit_fam not in ['célibataire']:
                    if nbr_enf <= 3:
                        return 5 + nbr_enf * 5
                    else:
                        return 20
                else:
                    return 0

            demandes_traiter_df['S'] = demandes_traiter_df.apply(calculate_S, axis=1)

            # Calculate D
            def calculate_D(row, historique_df):
                matricule = row['Matricule']
                matricule_rows = historique_df[historique_df['Matricule'] == matricule]
                total_D = 0

                for index, hist_row in matricule_rows.iterrows():
                    date_de_debut_sejour = hist_row['Date debut sejour']
                    year_difference = datetime.now().year - date_de_debut_sejour.year

                    if year_difference == 1:
                        total_D += 140
                    elif year_difference == 2:
                        total_D += 90
                    elif year_difference == 3:
                        total_D += 50
                    elif year_difference == 4:
                        total_D += 20

                return total_D

            demandes_traiter_df['D'] = demandes_traiter_df.apply(calculate_D, axis=1, historique_df=historique_df)
            demandes_traiter_df['D'] = demandes_traiter_df['D'].fillna(0)

            # Calculate P
            demandes_traiter_df['P'] = 2 * (demandes_traiter_df['A'] + demandes_traiter_df['S']) - demandes_traiter_df['D']

            # Sorting based on 'P'
            demandes_traiter_df = demandes_traiter_df.sort_values(by=['P', 'A', 'S', 'Date de la demande'], ascending=[False, False, False, False])

            # Save the DataFrame to the database
            save_to_database(demandes_traiter_df,request)
            messages.success(request, "Le traitement des fichiers a été effectué avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue : {str(e)}")

    return redirect('list_generated')  # Replace with the actual view name

def process_files(request):
    form = CalculForm(request.POST, request.FILES)
    if form.is_valid():
        print(" i was insideof process files ")
        # Retrieve date values from the form
        date_debut_sejour = form.cleaned_data.get('date_debut_sejour')
        date_fin_sejour = form.cleaned_data.get('date_fin_sejour')
        

        # Check if dates are provided
        if not date_debut_sejour or not date_fin_sejour:
            messages.error(request, "Les dates de début et de fin de séjour sont requises.")
            return redirect('list_generated')  # Replace with the actual view name
        
        try:
            # Fetch data from the models
            historique_queryset = Historique.objects.all().values()
            agents_queryset = Agent.objects.all().values()
            demandes_queryset = Demande.objects.all().values()
            # Convert the querysets to DataFrames
            historique_df = pd.DataFrame(list(historique_queryset))
            agents_df = pd.DataFrame(list(agents_queryset))
            demandes_df = pd.DataFrame(list(demandes_queryset))

            # Column name mappings
            historique_rename_dict = {
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'nombre_nuites': 'Nombre de nuites',
            }
            historique_df.rename(columns=historique_rename_dict, inplace=True)

            agents_rename_dict = {
                'matricule': 'matricule',
                'nom_prenom': 'nom_prenom',
                'date_naissance': 'date_naissance',
                'date_embauche': 'date_embauche',
                'nombre_enf': 'NOMBRE_ENF',
                'date_debut_retraite': 'date_debut_retraite',
            }
            agents_df.rename(columns=agents_rename_dict, inplace=True)

            demandes_rename_dict = {
                'numero_demande': 'N° Demande',
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'type_de_vue': 'Type de vue',
                'nombre_nuites': 'Nombre de nuites',
                'statut': 'Statut',
                'nature_periode': 'Nature Periode',
                'site': 'Site'
            }
            demandes_df.rename(columns=demandes_rename_dict, inplace=True)

            # Filter rows based on date_debut_sejour and date_fin_sejour
            demandes_df = demandes_df[
                (demandes_df['Date debut sejour'] == date_debut_sejour) &
                (demandes_df['Date fin sejour'] == date_fin_sejour)
            ]

            # Create a new DataFrame 'demandes_traiter' that contains the exact content of the 'demandes' file
            demandes_traiter_df = demandes_df.copy()

            # Step 4: Filter Rows
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Bloquée']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

            # Step 5: Add retraite Column and Calculate Its Values
            from datetime import datetime
            demandes_traiter_df = demandes_traiter_df.merge(
                agents_df[['matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF', 'date_debut_retraite']],
                left_on='Matricule', right_on='matricule', how='left'
            )

            # Convert dates to datetime
            demandes_traiter_df['date_embauche'] = pd.to_datetime(demandes_traiter_df['date_embauche'], format='%d/%m/%Y')
            demandes_traiter_df['Date debut sejour'] = pd.to_datetime(demandes_traiter_df['Date debut sejour'], format='%d/%m/%Y')
            demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')
            
            # Check for missing values in required columns of demandes_traiter_df
            required_columns = ['Matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF']
            missing_data_df = demandes_traiter_df[demandes_traiter_df[required_columns].isna().any(axis=1)]
            print("this is the messign data shape")
            print(missing_data_df.shape)


            if not missing_data_df.empty:
                # Pass missing data details to the template
                print("i was heeer")
                # Retrieve missing data from session
                request.session['missing_data'] = missing_data_df.to_json(orient='records')
                print(request.session['missing_data'])
                print("Missing data has been stored in the session.")
        
                return redirect('list_generated')  # Redirect to the page where you will handle the modal
               

            

            # Filter expired date_debut_retraite
            save_rejected_records(demandes_traiter_df, request)
            demandes_traiter_df = demandes_traiter_df[
                (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                (demandes_traiter_df['date_debut_retraite'].isna())
            ]

            if demandes_traiter_df.empty:
                messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                return redirect('list_generated')

          

            # Define the fixed date: June 1st of the current year
            fixed_date = datetime(datetime.now().year, 6, 1)

            # Calculate the difference in months
            demandes_traiter_df['A'] = demandes_traiter_df['date_embauche'].apply(
                lambda x: relativedelta(fixed_date, x).years * 12 + relativedelta(fixed_date, x).months
            )


            # Calculate S
            def calculate_S(row):
                sit_fam = row['sit_fam'].strip().lower()
                nbr_enf = row['NOMBRE_ENF']

                if sit_fam not in ['célibataire']:
                    if nbr_enf <= 3:
                        return 5 + nbr_enf * 5
                    else:
                        return 20
                else:
                    return 0

            demandes_traiter_df['S'] = demandes_traiter_df.apply(calculate_S, axis=1)

            # Calculate D
            def calculate_D(row, historique_df):
                matricule = row['Matricule']
                matricule_rows = historique_df[historique_df['Matricule'] == matricule]
                total_D = 0

                for index, hist_row in matricule_rows.iterrows():
                    date_de_debut_sejour = hist_row['Date debut sejour']
                    year_difference = datetime.now().year - date_de_debut_sejour.year

                    if year_difference == 1:
                        total_D += 140
                    elif year_difference == 2:
                        total_D += 90
                    elif year_difference == 3:
                        total_D += 50
                    elif year_difference == 4:
                        total_D += 20

                return total_D

            demandes_traiter_df['D'] = demandes_traiter_df.apply(calculate_D, axis=1, historique_df=historique_df)
            demandes_traiter_df['D'] = demandes_traiter_df['D'].fillna(0)

            # Calculate P
            demandes_traiter_df['P'] = 2 * (demandes_traiter_df['A'] + demandes_traiter_df['S']) - demandes_traiter_df['D']

            # Sorting based on 'P'
            demandes_traiter_df = demandes_traiter_df.sort_values(by=['P', 'A', 'S', 'Date de la demande'], ascending=[False, False, False, False])

            # Save the DataFrame to the database
            save_to_database(demandes_traiter_df,request)

            # Convert date strings to date objects
            date_debut_sejour = datetime.strptime(date_debut_sejour, '%Y-%m-%d').date()
            date_fin_sejour = datetime.strptime(date_fin_sejour, '%Y-%m-%d').date()

            # Store date objects in session as strings
            request.session['date_debut_sejour'] = date_debut_sejour.strftime('%Y-%m-%d')
            request.session['date_fin_sejour'] = date_fin_sejour.strftime('%Y-%m-%d')

            messages.success(request, "Le traitement des fichiers a été effectué avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue : {str(e)}")

    return redirect('list_generated')  # Replace with the actual view name

def proceed_without_storing(request):
    if request.method == 'POST' and request.FILES:
        # Récupérer les fichiers téléchargés
        agents_file = request.FILES['agents_file']
        historique_file = request.FILES['historique_file']
        demandes_file = request.FILES['demandes_file']
            
        # Lire les fichiers Excel
        agents_df = pd.read_excel(agents_file)
        historique_df = pd.read_excel(historique_file)
        demandes_df = pd.read_excel(demandes_file)

        # Retrieve dates directly from the POST request
        date_debut_sejour = request.POST.get('date_debut_sejour')
        date_fin_sejour = request.POST.get('date_fin_sejour')
        print(" here is the debug dates:")
        print(date_debut_sejour)
        print(date_fin_sejour)


        # Check if dates are provided
        if not date_debut_sejour or not date_fin_sejour:
            messages.error(request, "Les dates de début et de fin de séjour sont requises.")
            return redirect('list_generated')  # Replace with the actual view name
        
        try:
            # Filter rows based on date_debut_sejour and date_fin_sejour
            demandes_df = demandes_df[
                (demandes_df['Date debut sejour'] == date_debut_sejour) &
                (demandes_df['Date fin sejour'] == date_fin_sejour)
            ]

            # Create a new DataFrame 'demandes_traiter' that contains the exact content of the 'demandes' file
            demandes_traiter_df = demandes_df.copy()

            # Step 4: Filter Rows
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Bloquée']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

            # Step 5: Add retraite Column and Calculate Its Values
            from datetime import datetime
            demandes_traiter_df = demandes_traiter_df.merge(
                agents_df[['matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF', 'date_debut_retraite']],
                left_on='Matricule', right_on='matricule', how='left'
            )

            # Convert dates to datetime
            demandes_traiter_df['date_embauche'] = pd.to_datetime(demandes_traiter_df['date_embauche'], format='%d/%m/%Y')
            demandes_traiter_df['Date debut sejour'] = pd.to_datetime(demandes_traiter_df['Date debut sejour'], format='%d/%m/%Y')
            demandes_traiter_df['Date fin sejour'] = pd.to_datetime(demandes_traiter_df['Date fin sejour'], format='%d/%m/%Y')
            demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')
            
            # Check for missing values in required columns of demandes_traiter_df
            required_columns = ['Matricule', 'date_embauche', 'sit_fam', 'NOMBRE_ENF']
            
            demandes_traiter_df = demandes_traiter_df.dropna(subset=required_columns)
            messages.success(request, "L'opération a été effectuée avec succès après avoir ignoré les données manquantes. Les calculs ont été poursuivis avec les données disponibles.")
                

            # Filter expired date_debut_retraite
            save_rejected_records(demandes_traiter_df, request)

            demandes_traiter_df = demandes_traiter_df[
                (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                (demandes_traiter_df['date_debut_retraite'].isna())
            ]

            if demandes_traiter_df.empty:
                messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                return redirect('list_generated')

          

            # Define the fixed date: June 1st of the current year
            fixed_date = datetime(datetime.now().year, 6, 1)

            # Calculate the difference in months
            demandes_traiter_df['A'] = demandes_traiter_df['date_embauche'].apply(
                lambda x: relativedelta(fixed_date, x).years * 12 + relativedelta(fixed_date, x).months
            )


            # Calculate S
            def calculate_S(row):
                sit_fam = row['sit_fam'].strip().lower()
                nbr_enf = row['NOMBRE_ENF']

                if sit_fam not in ['célibataire']:
                    if nbr_enf <= 3:
                        return 5 + nbr_enf * 5
                    else:
                        return 20
                else:
                    return 0

            demandes_traiter_df['S'] = demandes_traiter_df.apply(calculate_S, axis=1)

            # Calculate D
            def calculate_D(row, historique_df):
                matricule = row['Matricule']
                matricule_rows = historique_df[historique_df['Matricule'] == matricule]
                total_D = 0

                for index, hist_row in matricule_rows.iterrows():
                    date_de_debut_sejour = hist_row['Date debut sejour']
                    year_difference = datetime.now().year - date_de_debut_sejour.year

                    if year_difference == 1:
                        total_D += 140
                    elif year_difference == 2:
                        total_D += 90
                    elif year_difference == 3:
                        total_D += 50
                    elif year_difference == 4:
                        total_D += 20

                return total_D

            demandes_traiter_df['D'] = demandes_traiter_df.apply(calculate_D, axis=1, historique_df=historique_df)
            demandes_traiter_df['D'] = demandes_traiter_df['D'].fillna(0)

            # Calculate P
            demandes_traiter_df['P'] = 2 * (demandes_traiter_df['A'] + demandes_traiter_df['S']) - demandes_traiter_df['D']

            # Sorting based on 'P'
            demandes_traiter_df = demandes_traiter_df.sort_values(by=['P', 'A', 'S', 'Date de la demande'], ascending=[False, False, False, False])

            # Save the DataFrame to the database
            save_to_database(demandes_traiter_df,request)
            messages.success(request, "Le traitement des fichiers a été effectué avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue : {str(e)}")

    return redirect('list_generated')  # Replace with the actual view name


from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def create_pdf(principal_data, waiting_data, col_widths, title_text, title_text_wait, filename='sorted_table.pdf'):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.black,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    normal_style = styles['Normal']
    normal_style.fontName = 'Helvetica'
    normal_style.fontSize = 10

    # Add title text as a Paragraph
    title_paragraph = Paragraph(title_text, title_style)
    elements.append(title_paragraph)
    elements.append(Spacer(1, 20))  # 20 points of space

    # Remove 'Type de vue' column from principal_data and waiting_data
    principal_data = [row for row in principal_data if row[0] != 'type_de_vue']
    waiting_data = [row for row in waiting_data if row[0] != 'type_de_vue']

    # Create principal table
    principal_table = Table(principal_data, colWidths=col_widths)
    principal_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font style (bold)
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Font size for header row
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data rows background color
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),  # Adjust font size if necessary
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),  # Alternating row colors
        ('COLBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.lightgrey])  # Alternating column colors
    ])
    principal_table.setStyle(principal_style)
    elements.append(principal_table)
    elements.append(Spacer(1, 24))

    # Add message if waiting list is empty
    if len(waiting_data) <= 1:  # Check if waiting_data has only the header row
        empty_waiting_list_text = f"*{title_text_wait} est vide"
        empty_waiting_list_paragraph = Paragraph(empty_waiting_list_text, normal_style)
        elements.append(Spacer(1, 20))  # 20 points of space
        elements.append(empty_waiting_list_paragraph)
    else:
        # Add page break and title for waiting list
        elements.append(PageBreak())
        waiting_title_paragraph = Paragraph(title_text_wait, title_style)
        elements.append(waiting_title_paragraph)
        elements.append(Spacer(1, 20))  # 20 points of space

        # Create waiting list table
        waiting_table = Table(waiting_data, colWidths=col_widths)
        waiting_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),  # Header background color
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font style (bold)
            ('FONTSIZE', (0, 0), (-1, 0), 10),  # Font size for header row
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data rows background color
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),  # Adjust font size if necessary
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),  # Alternating row colors
            ('COLBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.lightgrey])  # Alternating column colors
        ])
        waiting_table.setStyle(waiting_style)
        elements.append(waiting_table)

    # Build the PDF
    doc.build(elements)
    pdf_data = buffer.getvalue()
    buffer.close()

    return pdf_data

def quota_list(request):
    quotas = Quota.objects.all()
    type_de_vue_list = Quota.objects.values_list('type_de_vue', flat=True).distinct()
    return render(request, 'listings/quota.html', {'quotas': quotas, 'type_de_vue_list': type_de_vue_list})



def add_quota(request):
    if request.method == 'POST':
        ville = request.POST.get('ville')
        type_de_vue = request.POST.get('type_de_vue')
        quota_value = request.POST.get('quota_value')
        
        if Quota.objects.filter(ville=ville, type_de_vue=type_de_vue).exists():
            messages.error(request, 'Un quota pour cette ville et ce type de vue existe déjà.')
        else:
            Quota.objects.create(ville=ville, type_de_vue=type_de_vue, quota_value=quota_value)
            messages.success(request, 'Quota ajouté avec succès.')

    return redirect('quota_list')  # Redirect to the view showing the table

def delete_quota(request):
    if request.method == 'POST':
        ville = request.POST.get('ville')
        type_de_vue = request.POST.get('type_de_vue')
        
        deleted_count, _ = Quota.objects.filter(ville=ville, type_de_vue=type_de_vue).delete()
        if deleted_count > 0:
            messages.success(request, 'Quota supprimé avec succès.')
        else:
            messages.error(request, 'Aucun quota trouvé pour la ville et le type de vue spécifiés.')

    return redirect('quota_list')  # Redirect to the view showing the table

def update_quota(request):
    if request.method == 'POST':
        ville = request.POST.get('original_ville')
        type_de_vue = request.POST.get('original_type_de_vue')
        new_quota_value = request.POST.get('quota_value')
        
        try:
            quota = Quota.objects.get(ville=ville, type_de_vue=type_de_vue)
            quota.quota_value = new_quota_value
            quota.save()
            messages.success(request, 'Quota mis à jour avec succès.')
        except Quota.DoesNotExist:
            messages.error(request, 'Quota non trouvé pour la ville et le type de vue spécifiés.')

    return redirect('quota_list')  # Redirect to the view showing the table


from django.shortcuts import render, redirect
from django.contrib import messages
import pandas as pd
from .models import Historique


def historique(request):
    # Display the page with the table and form
    historiques = Historique.objects.all()
    return render(request, 'listings/historique.html', {'historiques': historiques})
import pandas as pd
from django.shortcuts import redirect
from django.contrib import messages
from .models import Historique
import logging

logger = logging.getLogger(__name__)

def upload_excel_historique(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, "Aucun fichier Excel trouvé.")
            return redirect('historique')

        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier Excel: {e}")
            return redirect('historique')

        # Clean the table before uploading new data
        Historique.objects.all().delete()

        # Validate columns
        required_columns = ['Site', 'Nom agent', 'Prenom agent', 'Matricule',
                            'Date de la demande', 'Date debut sejour', 'Date fin sejour',
                            'Type de vue', 'Nombre de nuites']

        if not all(column in df.columns for column in required_columns):
            messages.error(request, "Le fichier Excel n'a pas les colonnes requises.")
            return redirect('historique')

        # Track rows with missing values in required columns
        rows_with_missing_values = df[df[required_columns].isnull().any(axis=1)]
        df = df.dropna(subset=required_columns)  # Drop rows with missing values in required columns

        # Insert data into the database
        inserted_rows = 0
        for _, row in df.iterrows():
            try:
                Historique.objects.create(
                    ville=row['Site'],
                    nom_agent=row['Nom agent'],
                    prenom_agent=row['Prenom agent'],
                    matricule=row['Matricule'],
                    date_demande=row['Date de la demande'],
                    date_debut_sejour=row['Date debut sejour'],
                    date_fin_sejour=row['Date fin sejour'],
                    type_de_vue=row['Type de vue'],
                    nombre_nuites=row['Nombre de nuites'],
                )
                inserted_rows += 1
            except Exception as e:
                logger.error(f"Erreur lors de l'insertion de la ligne {row}: {e}")

        if not rows_with_missing_values.empty:
            messages.warning(request, "Certaines lignes avaient des valeurs manquantes et n'ont pas été insérées.", extra_tags='missing-values')
        else:
            messages.success(request, f"Fichier Excel traité avec succès. {inserted_rows} lignes insérées.")

    return redirect('historique')

def delete_historique(request):
    if request.method == 'POST':
        Historique.objects.all().delete()
        messages.success(request, "Toutes les données ont été supprimées.")
    return redirect('historique')

# login views

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from .forms import SignUpForm
from .models import Profile
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Profile
from .forms import SignUpForm

from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import SignUpForm
from .models import Profile

def login_page(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            profile_exists = Profile.check_profile(username, password)
            if profile_exists:
                profile = Profile.objects.get(username=username, password=password)
                request.session['profile_id'] = profile.id
                request.session.set_expiry(1800)  # Set session to expire in 30 minutes
                context = {
                  'profile': profile,
                  }
                messages.success(request, 'Connexion réussie ! Bienvenue sur votre profil.')
                return render(request, 'listings/profile.html', context)
            else:
                messages.error(request, 'Échec de la connexion. Assurez-vous que vos identifiants sont corrects.')
        else:
            messages.error(request, 'Erreur dans le formulaire. Veuillez vérifier vos données.')
    else:
        form = SignUpForm()
    
    return render(request, 'listings/signup.html', {'form': form})

    





def profile_page(request):
    profile_id = request.session.get('profile_id')
    print(" i am inside of profie view ")
    if profile_id:
        profile = Profile.objects.get(id=profile_id)
        context = {
                'profile': profile,
            }
        return render(request, 'listings/profile.html', context)
    else:
        messages.info(request, "Erreur de session. Veuillez vous reconnecter.")
        return render(request, 'listings/signup.html')

from django.shortcuts import redirect
from django.contrib import messages

def logout_view(request):
    # Clear the profile_id from the session to log the user out
    if 'profile_id' in request.session:
        del request.session['profile_id']
        messages.success(request, 'Vous avez été déconnecté avec succès.')
    else:
        messages.info(request, 'Vous n\'êtes pas connecté.')

    # Redirect the user to the login page or home page
    return redirect('signup')  # Change 'signup' to your login or home view name if necessary



from django.shortcuts import get_object_or_404, render, redirect
from .forms import Agent, AgentForm, Demande, DemandeForm
from django.db.models import Count
from django.http import HttpResponseBadRequest
import pandas as pd
import json

from .models import DemandesTraiter



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Agent
from .forms import AgentForm

def AgentsAffichage(request):
    if request.method == 'POST':
        form = AgentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Agent ajouté avec succès !")
            return redirect('AgentsAffichage')
    else:
        form = AgentForm()

    agents = Agent.objects.all()
    return render(request, 'listings/agentsCrud.html', {'data': agents, 'form': form})


def edit_agent(request, agent_id):
    agent = get_object_or_404(Agent, pk=agent_id)
    if request.method == 'POST':
        form = AgentForm(request.POST, instance=agent)
        if form.is_valid():
            form.save()
            messages.success(request, "Agent modifié avec succès !")
            return redirect('AgentsAffichage')
    else:
        form = AgentForm(instance=agent)
    return render(request, 'listings/edit_agent.html', {'form': form})


def delete_agent(request, agent_id):
    if request.method == 'POST':
        agent = get_object_or_404(Agent, id=agent_id)
        agent.delete()
        messages.success(request, "Agent supprimé avec succès.")
    else:
        messages.error(request, "La suppression a échoué.")
    return redirect('AgentsAffichage')


import pandas as pd
from django.contrib import messages
from django.shortcuts import redirect
from .models import Agent

import pandas as pd
from django.contrib import messages
from django.shortcuts import redirect
from .models import Agent

from django.shortcuts import redirect
from django.contrib import messages
import pandas as pd
from .models import Agent

def upload_excel_agents(request):
    if request.method == 'POST':
        if 'excelFile' not in request.FILES:
            messages.error(request, "Aucun fichier n'a été téléchargé.")
            return redirect('AgentsAffichage')

        excel_file = request.FILES['excelFile']
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, "Format de fichier invalide. Veuillez télécharger un fichier .xlsx.")
            return redirect('AgentsAffichage')

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier Excel : {str(e)}")
            return redirect('AgentsAffichage')

        expected_columns = ['matricule', 'nom_prenom', 'date_naissance', 'sit_fam', 'date_embauche', 'nombre_enf', 'date_debut_retraite']
        df.columns = df.columns.str.lower()
        if not all(col in df.columns for col in expected_columns):
            messages.error(request, "Format de fichier Excel invalide. Les colonnes attendues sont manquantes.")
            return redirect('AgentsAffichage')

        # Delete all existing records from the Agent table
        try:
            Agent.objects.all().delete()
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression des anciens enregistrements : {str(e)}")
            return redirect('AgentsAffichage')

        for index, row in df.iterrows():
            if pd.isnull(row['matricule']) or pd.isnull(row['nom_prenom']) or pd.isnull(row['date_naissance']):
                messages.error(request, f"Données requises manquantes à la ligne {index + 1}.")
                return redirect('AgentsAffichage')
            try:
                agent = Agent(
                    matricule=row['matricule'],
                    nom_prenom=row['nom_prenom'],
                    date_naissance=row['date_naissance'] if not pd.isna(row['date_naissance']) else None,
                    sit_fam=row['sit_fam'],
                    date_embauche=row['date_embauche'] if not pd.isna(row['date_embauche']) else None,
                    nombre_enf=row['nombre_enf'],
                    date_debut_retraite=row.get('date_debut_retraite') if not pd.isna(row.get('date_debut_retraite')) else None
                )
                agent.save()
            except Exception as e:
                messages.error(request, f"Erreur lors de l'enregistrement des données : {str(e)}")
                return redirect('AgentsAffichage')

        messages.success(request, "Fichier téléchargé et traité avec succès !")
        return redirect('AgentsAffichage')
    else:
        return redirect('AgentsAffichage')



def delete_all_agents(request):
    if request.method == 'POST':
        Agent.objects.all().delete()
        messages.success(request, "Tous les agents ont été supprimés avec succès !")
        return redirect('AgentsAffichage')

    return redirect('AgentsAffichage')




###################Demande View####################################
from django.shortcuts import render
from django.db.models import Count
import json
from .models import Agent, Demande, DemandesTraiter

def statistiques(request):
    try:
        total_agents = Agent.objects.count()
        total_demandes = Demande.objects.count()

        demandes_par_vue = list(Demande.objects.values('type_de_vue').annotate(count=Count('type_de_vue')))
        demandes_par_ville = list(Demande.objects.values('ville').annotate(count=Count('ville')))

        nbr_vue = Demande.objects.values('type_de_vue').distinct().count()

        context = {
            'total_agents': total_agents,
            'total_demandes': total_demandes,
            'demandes_par_vue': json.dumps(demandes_par_vue),
            'demandes_par_ville': json.dumps(demandes_par_ville),
            'nbr_vue': nbr_vue,
            'demandes': DemandesTraiter.objects.all(),
            'rejected_demandes': RejectedDemandesRetrait.objects.all(),  # Add this line
        }

        return render(request, 'listings/dashboard.html', context)

    except OperationalError as e:
        # Handle database operational errors
        messages.error(request, "Une erreur de base de données s'est produite : {}".format(str(e)))
        return render(request, 'listings/dashboard.html', {'error': str(e)})

    except Exception as e:
        # Handle any other exceptions
        messages.error(request, "Une erreur inattendue s'est produite : {}".format(str(e)))
        return render(request, 'listings/dashboard.html', {'error': str(e)})

import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from .models import Demande
from .forms import DemandeForm

# View for uploading Excel and inserting data
import pandas as pd
from django.shortcuts import redirect, render
from django.contrib import messages
from .models import Demande

def upload_excel_demandes(request):
    if request.method == 'POST':
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            try:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(excel_file)

                # Check if required columns exist in the DataFrame
                required_columns = [
                    'Ville',
                    'Nom agent',
                    'Prenom agent',
                    'Matricule',
                    'Date de la demande',
                    'Date debut sejour',
                    'Date fin sejour',
                    'Type de vue',
                    'Nombre de nuites',
                    'Statut',
                    'Nature Periode',
                    'Site',
                    'N° Demande'  # Added new column here
                ]

                # Ensure all required columns are present
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f'Colonne(s) manquante(s) dans le fichier Excel: {", ".join(missing_columns)}')
                    return redirect('demande_list')

                # Delete all existing records from the Demande table
                Demande.objects.all().delete()

                # Iterate over each row in the DataFrame and create Demande instances
                for _, row in df.iterrows():
                    Demande.objects.create(
                        ville=row.get('Ville', ''),
                        nom_agent=row.get('Nom agent', ''),
                        prenom_agent=row.get('Prenom agent', ''),
                        matricule=row.get('Matricule', ''),
                        date_demande=row.get('Date de la demande', pd.NaT),
                        date_debut_sejour=row.get('Date debut sejour', pd.NaT),
                        date_fin_sejour=row.get('Date fin sejour', pd.NaT),
                        type_de_vue=row.get('Type de vue', 'Inconnu'),
                        nombre_nuites=row.get('Nombre de nuites', 0),
                        statut=row.get('Statut', ''),
                        nature_periode=row.get('Nature Periode', ''),
                        site=row.get('Site', ''),
                        numero_demande=row.get('N° Demande', '')  # Handle new column here
                    )

                messages.success(request, 'Fichier Excel téléchargé et données insérées avec succès.')
            except Exception as e:
                messages.error(request, f'Erreur lors du traitement du fichier Excel: {e}')
            
            return redirect('demande_list')
    
    return render(request, 'listings/demandesCrud.html')


# View for listing demandes
def demande_list(request):
    demandes = Demande.objects.all()
    return render(request, 'listings/demandesCrud.html', {'demandes': demandes})


# View for adding a new demande
# View for adding a demande
def add_demande(request):
    if request.method == 'POST':
        form = DemandeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Demande ajoutée avec succès.')
            return redirect('demande_list')
        else:
            messages.error(request, 'Erreur lors de l\'ajout de la demande. Veuillez vérifier les champs.')
    else:
        form = DemandeForm()
    return render(request, 'listings/demandesCrud.html', {'form': form})

# View for editing a demande
def edit_demande(request, demande_id):
    demande = get_object_or_404(Demande, id=demande_id)
    
    if request.method == 'POST':
        form = DemandeForm(request.POST, instance=demande)
        if form.is_valid():
            form.save()
            messages.success(request, 'Demande modifiée avec succès.')
            return redirect('demande_list')
        else:
            messages.error(request, 'Erreur lors de la modification de la demande. Veuillez vérifier les champs.')
            print("Form errors:", form.errors)  # Debug: Print form errors to console
    else:
        form = DemandeForm(instance=demande)
    
    print("Form data:", form.initial)  # Debug: Print initial form data
    return redirect('demande_list')


# View for deleting a demande
def delete_demande(request, demande_id):
    if request.method == 'POST':
        try:
            demande = get_object_or_404(Demande, pk=demande_id)
            demande.delete()
            messages.success(request, 'Demande supprimée avec succès.')
        except Demande.DoesNotExist:
            messages.error(request, 'Demande non trouvée.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression de la demande: {e}')
        return redirect('demande_list')
    messages.error(request, 'Erreur lors de la suppression de la demande.')
    return redirect('demande_list')

# View for deleting all demandes
def delete_all_demandes(request):
    if request.method == 'POST':
        try:
            Demande.objects.all().delete()
            messages.success(request, 'Toutes les demandes ont été supprimées avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression des demandes: {e}')
        return redirect('demande_list')
    messages.error(request, 'Erreur lors de la suppression des demandes.')
    return redirect('demande_list')

from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from .models import Demande

def get_demande_data(request, demande_id):
    demande = get_object_or_404(Demande, id=demande_id)
    data = {
        'ville': demande.ville,
        'nom_agent': demande.nom_agent,
        'prenom_agent': demande.prenom_agent,
        'matricule': demande.matricule,
        'date_demande': demande.date_demande,
        'date_debut_sejour': demande.date_debut_sejour,
        'date_fin_sejour': demande.date_fin_sejour,
        'type_de_vue': demande.type_de_vue,
        'nombre_nuites': demande.nombre_nuites,
    }
    return JsonResponse(data)


import pandas as pd
from django.http import HttpResponse
from .models import RejectedDemandesRetrait

def download_rejected_agents(request):
    # Query rejected demands from the database
    rejected_agents = RejectedDemandesRetrait.objects.all()

    # Create a DataFrame from the queried data
    df = pd.DataFrame(list(rejected_agents.values(
        'nom_agent', 'prenom_agent', 'matricule'
    )))

    # Create an HTTP response with CSV content
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rejected_agents.csv"'

    # Convert DataFrame to CSV and write to response
    df.to_csv(path_or_buf=response, index=False)

    return response


# listings/views.py
from django.shortcuts import redirect
from django.contrib import messages
from .models import DemandesTraiter, RejectedDemandesRetrait
from django.views.decorators.http import require_POST

@require_POST
def delete_all_demandes_traiter(request):
    try:
        DemandesTraiter.objects.all().delete()
        messages.success(request, 'Toutes les demandes traitées ont été supprimées avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression des demandes traitées: {e}')
    return redirect('dashboard')  # Redirect to the page displaying the statistics

@require_POST
def delete_all_demandes_rejetees(request):
    try:
        RejectedDemandesRetrait.objects.all().delete()
        messages.success(request, 'Toutes les demandes rejetées ont été supprimées avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression des demandes rejetées: {e}')
    return redirect('dashboard')  # Redirect to the page displaying the statistics

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
import io
from .models import RejectedDemandesRetrait  # Import the correct model

def download_pdf_rejected_demandes(request):
    rejected_demandes = RejectedDemandesRetrait.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading2']
    body_style = styles['BodyText']

    # Title
    title = Paragraph("Liste des Demandes Rejetées", title_style)
    elements.append(title)

    # Table header
    data = [
        ["Numéro Demande", "Nom Agent", "Prénom Agent", "Date Début Séjour", "Date Fin Séjour"]
    ]
    
    # Add data
    for demande in rejected_demandes:
        data.append([
            demande.numero_demande,
            demande.nom_agent,
            demande.prenom_agent,
            demande.date_debut_sejour.strftime("%d-%m-%Y"),
            demande.date_fin_sejour.strftime("%d-%m-%Y")
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f5f5f5'),
        ('GRID', (0, 0), (-1, -1), 1, '#000000'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="demandes_rejetees.pdf"'
    return response
    
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
import io
from .models import DemandesTraiter  # Import the correct model

def download_pdf_demandes_traiter(request):
    demandes = DemandesTraiter.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading2']
    body_style = styles['BodyText']

    # Title
    title = Paragraph("Liste des Demandes Traitées", title_style)
    elements.append(title)

    # Table header
    data = [
        ["Numéro Demande", "Ville", "Matricule", "Date Début Séjour", "Date Fin Séjour", "A", "D", "S", "P"]
    ]
    
    # Add data
    for demande in demandes:
        data.append([
            demande.numero_demande,
            demande.ville,
            #demande.nom_agent,
           # demande.prenom_agent,
            demande.matricule,
            demande.date_debut_sejour.strftime("%d-%m-%Y"),
            demande.date_fin_sejour.strftime("%d-%m-%Y"),
            demande.A,
            demande.D,
            demande.S,
            demande.P
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f5f5f5'),
        ('GRID', (0, 0), (-1, -1), 1, '#000000'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="demandes_traiter.pdf"'
    return response


import pandas as pd
from django.http import HttpResponse

def download_excel(request, ville, type_de_vue):
    # Retrieve data from the database
    data = get_data_from_database(ville, type_de_vue)
    df = pd.DataFrame(data)
    
    # Print column names to debug
    print("DataFrame columns:", df.columns.tolist())
    
    # Sort the DataFrame based on 'P' in descending order
    sorted_table = df.sort_values(by=['P'], ascending=[False]).reset_index(drop=True)
    
    # Define columns of interest based on the model
    # Include all columns from the model
    columns_of_interest = [
        'numero_demande', 'ville', 'nom_agent', 'prenom_agent', 'matricule',
        'date_debut_sejour', 'date_fin_sejour', 'type_de_vue', 'A', 'D', 'S', 'P'
    ]
    
    # Check if all columns exist in the DataFrame
    missing_cols = [col for col in columns_of_interest if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns: {missing_cols}")

    # Prepare the selected data
    selected_data = sorted_table[columns_of_interest].copy()
    selected_data['date_debut_sejour'] = pd.to_datetime(selected_data['date_debut_sejour']).dt.strftime('%Y-%m-%d')
    selected_data['date_fin_sejour'] = pd.to_datetime(selected_data['date_fin_sejour']).dt.strftime('%Y-%m-%d')

    # Create an Excel writer object and save the DataFrame to an Excel file in memory
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{ville}_{type_de_vue}.xlsx"'
    
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        selected_data.to_excel(writer, sheet_name='Liste', index=False)
        # Adjust column width (optional)
        for column in selected_data:
            col_idx = selected_data.columns.get_loc(column)
            writer.sheets['Liste'].set_column(col_idx, col_idx, max(selected_data[column].astype(str).map(len).max(), len(column)))

    return response
def download_all_demandes_excel(request):
    # Retrieve all rows from the DemandesTraiter table
    demandes = DemandesTraiter.objects.all()

    # Convert the queryset to a DataFrame
    data = list(demandes.values())
    df = pd.DataFrame(data)

    # Define the filename in French
    filename = 'toutes_les_demandes_traitées.xlsx'

    # Create an HttpResponse object with the appropriate headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Use pandas to create an Excel file in memory
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='DemandesTraiter', index=False)
        # Adjust column width (optional)
        for column in df:
            col_idx = df.columns.get_loc(column)
            writer.sheets['DemandesTraiter'].set_column(col_idx, col_idx, max(df[column].astype(str).map(len).max(), len(column)))

    return response


def download_all_rejected_demandes_excel(request):
    # Retrieve all rejected demandes
    rejected_demandes = RejectedDemandesRetrait.objects.all()
    df = pd.DataFrame(list(rejected_demandes.values()))

    # Define columns of interest based on the model
    columns_of_interest = ['numero_demande', 'ville', 'nom_agent', 'prenom_agent', 'matricule',
                           'date_debut_sejour', 'date_fin_sejour', 'type_de_vue', 'date_debut_retraite', 'date_de_la_demande']

    # Prepare the selected data
    selected_data = df[columns_of_interest].copy()
    selected_data['date_debut_sejour'] = pd.to_datetime(selected_data['date_debut_sejour']).dt.strftime('%Y-%m-%d')
    selected_data['date_fin_sejour'] = pd.to_datetime(selected_data['date_fin_sejour']).dt.strftime('%Y-%m-%d')
    selected_data['date_debut_retraite'] = pd.to_datetime(selected_data['date_debut_retraite']).dt.strftime('%Y-%m-%d')
    selected_data['date_de_la_demande'] = pd.to_datetime(selected_data['date_de_la_demande']).dt.strftime('%Y-%m-%d')

    # Create an Excel writer object and save the DataFrame to an Excel file in memory
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="toutes_Demandes_Rejetees.xlsx"'
    
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        selected_data.to_excel(writer, sheet_name='Liste', index=False)
        # Adjust column width (optional)
        for column in selected_data:
            col_idx = selected_data.columns.get_loc(column)
            writer.sheets['Liste'].set_column(col_idx, col_idx, max(selected_data[column].astype(str).map(len).max(), len(column)))

    return response



#################Periode Libre##########################

def ListsLibre(request):
    # Your logic for this view
    return render(request, 'listings/listesLibres.html')

def filter_and_rank_agents(request):
    if request.method == 'POST':
        form = CalculForm(request.POST)
        if form.is_valid():
            date_debut_sejour = form.cleaned_data['date_debut_sejour']
            date_fin_sejour = form.cleaned_data['date_fin_sejour']
            
            try:

                historique_queryset = Historique.objects.all().values()
                agents_queryset = Agent.objects.all().values()
                demandes_queryset = Demande.objects.all().values()
                
                # Convert the querysets to DataFrames
                historique_df = pd.DataFrame(list(historique_queryset))
                agents_df = pd.DataFrame(list(agents_queryset))
                demandes_df = pd.DataFrame(list(demandes_queryset))

                # Rename columns for consistency
                historique_df.rename(columns={
                    'ville': 'Ville',
                    'nom_agent': 'Nom agent',
                    'prenom_agent': 'Prenom agent',
                    'matricule': 'Matricule',
                    'date_demande': 'Date de la demande',
                    'date_debut_sejour': 'Date debut sejour',
                    'date_fin_sejour': 'Date fin sejour',
                    'nombre_nuites': 'Nombre de nuites'
                }, inplace=True)

                agents_df.rename(columns={
                    'matricule': 'Matricule',
                    'nom_prenom': 'Nom Prenom',
                    'date_naissance': 'Date Naissance',
                    'date_embauche': 'Date Embauche',
                    'nombre_enf': 'Nombre Enfants',
                    'date_debut_retraite': 'date_debut_retraite'
                }, inplace=True)

                demandes_df.rename(columns={
                    'numero_demande': 'Numero Demande',
                    'ville': 'Ville',
                    'nom_agent': 'Nom agent',
                    'prenom_agent': 'Prenom agent',
                    'matricule': 'Matricule',
                    'date_demande': 'Date de la demande',
                    'date_debut_sejour': 'Date debut sejour',
                    'date_fin_sejour': 'Date fin sejour',
                    'type_de_vue': 'Type de vue',
                    'nombre_nuites': 'Nombre de nuites',
                    'statut': 'Statut',
                    'nature_periode': 'Nature Periode',
                    'site': 'Site'
                }, inplace=True)

                # Filter demandes based on provided date range
                demandes_df = demandes_df[
                    (demandes_df['Date debut sejour'] >= date_debut_sejour) &
                    (demandes_df['Date fin sejour'] <= date_fin_sejour)
                ]
               
                # Additional filtering
                demandes_traiter_df = demandes_df.copy()
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Libre']
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

                # Convert date columns to datetime
                agents_df['Date Embauche'] = pd.to_datetime(agents_df['Date Embauche'], format='%d/%m/%Y')
                agents_df['Date Naissance'] = pd.to_datetime(agents_df['Date Naissance'], errors='coerce')
                historique_df['Date debut sejour'] = pd.to_datetime(historique_df['Date debut sejour'], errors='coerce')
                historique_df['Date fin sejour'] = pd.to_datetime(historique_df['Date fin sejour'], errors='coerce')

                

                # Process and merge data
                demandes_traiter_df = demandes_traiter_df.merge(
                    agents_df[['Matricule', 'Date Embauche', 'Nombre Enfants', 'Date Naissance','date_debut_retraite']],
                    left_on='Matricule', right_on='Matricule', how='left'
                )
                demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')

                # Filter expired date_debut_retraite
                save_rejected_records(demandes_traiter_df, request)
                demandes_traiter_df = demandes_traiter_df[
                    (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                    (demandes_traiter_df['date_debut_retraite'].isna())
                ]

                if demandes_traiter_df.empty:
                  messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                  return redirect('libre_generated')
            
                # Calculate additional fields
                def calculate_anciennete(date_embauche):
                    return relativedelta(datetime.now(), date_embauche).years * 12 + relativedelta(datetime.now(), date_embauche).months

                def calculate_age(date_naissance):
                    return relativedelta(datetime.now(), date_naissance).years
                print("first")
                print(demandes_traiter_df[['Date Embauche', 'Date Naissance']].isnull().sum())

                demandes_traiter_df['Anciennete'] = demandes_traiter_df['Date Embauche'].apply(calculate_anciennete)
                demandes_traiter_df['Age'] = demandes_traiter_df['Date Naissance'].apply(calculate_age)
                print("second")
                def calculate_sejour_count(matricule, historique_df):
                    return len(historique_df[historique_df['Matricule'] == matricule])

                def calculate_last_sejour(matricule, historique_df):
                    agent_hist = historique_df[historique_df['Matricule'] == matricule]
                    if agent_hist.empty:
                        return pd.NaT
                    return agent_hist['Date debut sejour'].max()

                demandes_traiter_df['Nombre Sejours'] = demandes_traiter_df['Matricule'].apply(lambda x: calculate_sejour_count(x, historique_df))
                demandes_traiter_df['Dernier Sejour'] = demandes_traiter_df['Matricule'].apply(lambda x: calculate_last_sejour(x, historique_df))

                # Handle NaT values in 'Dernier Sejour' column
                demandes_traiter_df['Dernier Sejour'] = demandes_traiter_df['Dernier Sejour'].fillna(pd.Timestamp.min)

                # Sorting agents according to the specified criteria
                def custom_sort(row):
                    if row['Nombre Sejours'] == 0:
                        return (0, -row['Anciennete'], -row['Age'], -row['Nombre Enfants'])
                    else:
                        return (
                            1,  # Priorité aux agents avec des séjours
                           row['Nombre Sejours'],  # Tri par nombre de séjours
                            row['Dernier Sejour'],  # Dernier séjour le plus loin en premier
                            #-row['Anciennete'],  # Ancienneté en mois
                            row['Date Embauche'],
                            -row['Age'],  # Âge
                            -row['Nombre Enfants']  # Nombre d'enfants
                        )

                demandes_traiter_df['Sort Key'] = demandes_traiter_df.apply(custom_sort, axis=1)
                demandes_traiter_df.sort_values(by='Sort Key', inplace=True)
                demandes_traiter_df.drop(columns='Sort Key', inplace=True)
                save_rejected_records(demandes_traiter_df, request)
                # Update AgentsLibre with correct Dernier Sejour
                AgentsLibre.objects.all().delete()  # Clean up old records
                for index, row in demandes_traiter_df.iterrows():
                    AgentsLibre.objects.create(
                        numero_demande=row['Numero Demande'],
                        ville=row['Ville'],
                        nom_agent=row['Nom agent'],
                        prenom_agent=row['Prenom agent'],
                        matricule=row['Matricule'],
                        date_debut_sejour=row['Date debut sejour'],
                        date_fin_sejour=row['Date fin sejour'],
                        type_de_vue=row['Type de vue'],
                        nombre_enfants=row['Nombre Enfants'],
                        age=row['Age'],
                        date_embauche=row['Date Embauche'],
                        anciennete=row['Anciennete'],
                        nombre_sejour=row['Nombre Sejours'],
                        dernier_sejour=row['Dernier Sejour']
                    )

                messages.success(request, "Les données ont été traitées et sauvegardées avec succès.")
                return redirect('libre_generated')
            except Exception as e:
                messages.error(request, f"Une erreur est survenue: {str(e)}")
                return redirect('libres_files')
    else:
        form = CalculForm()

    return redirect('libre_generated')








def libre_generated(request):
    # Fetch distinct combinations of 'ville' and 'type_de_vue'
    form = AffectationForm(request.POST)
    data = AgentsLibre.objects.values('ville', 'type_de_vue').distinct()
    
    # Group data by 'ville' and store unique 'type_de_vue' values

    grouped_data = defaultdict(set)
    for item in data:
        grouped_data[item['ville']].add(item['type_de_vue'])
    
    # Convert set to list for rendering
    context = {
        'data': {ville: list(types) for ville, types in grouped_data.items()},
        'demandesLibre': AgentsLibre.objects.all(),
        'rejected_demandes': RejectedDemandesRetrait.objects.all(),
        'form': form,
    }
    
    
    return render(request, 'listings/listesLibres.html', context)



from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from .models import AgentsLibre

import pandas as pd
from django.http import HttpResponse
from .models import AgentsLibre

def download_excel_libre(request, ville, type_de_vue):
    # Filter the AgentsLibre model based on the parameters
    filtered_agents = AgentsLibre.objects.filter(ville=ville, type_de_vue=type_de_vue)

    # Convert the queryset to a DataFrame
    df = pd.DataFrame(list(filtered_agents.values()))

    
    if df.empty:
        # Si le DataFrame est vide, créez un fichier Excel vide ou une feuille de calcul avec des messages appropriés
        wb = Workbook()
        ws = wb.active
        ws.append(['Aucune donnée disponible'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=rejected_demandes.xlsx'
        wb.save(response)
        return response

    # Define columns of interest
    columns_of_interest = ['numero_demande',  'matricule', 'nom_agent', 'prenom_agent','nombre_sejour','dernier_sejour', 'anciennete', date_embauche, 'age', 'nombre_enfants', 
                           'ville','type_de_vue','date_debut_sejour', 'date_fin_sejour']
    df = df[columns_of_interest]


    # Create a BytesIO buffer to hold the Excel data
    excel_buffer = BytesIO()
    
    # Write the DataFrame to the buffer
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    
    # Seek to the beginning of the stream
    excel_buffer.seek(0)

    # Return the Excel file as a downloadable response
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{ville}_{type_de_vue}.xlsx"'
    
    return response



############################################

def process_libres(request):
    form = CalculForm(request.POST, request.FILES)
    if form.is_valid():
        # Retrieve date values from the form
        date_debut_sejour = form.cleaned_data.get('date_debut_sejour')
        date_fin_sejour = form.cleaned_data.get('date_fin_sejour')

        # Check if dates are provided
        if not date_debut_sejour or not date_fin_sejour:
            messages.error(request, "Les dates de début et de fin de séjour sont requises.")
            return redirect('list_generated')  # Replace with the actual view name
        
        try:
            # Fetch data from the models
            historique_queryset = Historique.objects.all().values()
            agents_queryset = Agent.objects.all().values()
            demandes_queryset = Demande.objects.all().values()
            # Convert the querysets to DataFrames
            historique_df = pd.DataFrame(list(historique_queryset))
            agents_df = pd.DataFrame(list(agents_queryset))
            demandes_df = pd.DataFrame(list(demandes_queryset))

            # Column name mappings
            historique_rename_dict = {
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'nombre_nuites': 'Nombre de nuites',
            }
            historique_df.rename(columns=historique_rename_dict, inplace=True)

            agents_rename_dict = {
                'matricule': 'matricule',
                'nom_prenom': 'nom_prenom',
                'date_naissance': 'date_naissance',
                'date_embauche': 'date_embauche',
                'nombre_enf': 'NOMBRE_ENF',
                'date_debut_retraite': 'date_debut_retraite',
            }
            agents_df.rename(columns=agents_rename_dict, inplace=True)

            demandes_rename_dict = {
                'numero_demande': 'N° Demande',
                'ville': 'Ville',
                'nom_agent': 'Nom agent',
                'prenom_agent': 'Prenom agent',
                'matricule': 'Matricule',
                'date_demande': 'Date de la demande',
                'date_debut_sejour': 'Date debut sejour',
                'date_fin_sejour': 'Date fin sejour',
                'type_de_vue': 'Type de vue',
                'nombre_nuites': 'Nombre de nuites',
                'statut': 'Statut',
                'nature_periode': 'Nature Periode',
                'site': 'Site'
            }
            demandes_df.rename(columns=demandes_rename_dict, inplace=True)

            # Filter rows based on date_debut_sejour and date_fin_sejour
            demandes_df = demandes_df[
                (demandes_df['Date debut sejour'] == date_debut_sejour) &
                (demandes_df['Date fin sejour'] == date_fin_sejour)
            ]

            # Create a new DataFrame 'demandes_traiter' that contains the exact content of the 'demandes' file
            demandes_traiter_df = demandes_df.copy()

            # Step 4: Filter Rows
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Bloquée']
            demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

            # Step 5: Add retraite Column and Calculate Its Values
           
            demandes_traiter_df = demandes_df.merge(
                    historique_df[['Matricule', 'Date Embauche', 'Nombre Enfants', 'Date Debut Retraite', 'Date Naissance']],
                    on='Matricule',
                    how='left'
                )

            # Convert dates to datetime
            demandes_traiter_df['date_embauche'] = pd.to_datetime(demandes_traiter_df['date_embauche'], format='%d/%m/%Y')
            demandes_traiter_df['Date debut sejour'] = pd.to_datetime(demandes_traiter_df['Date debut sejour'], format='%d/%m/%Y')
            demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')

            # Filter expired date_debut_retraite
            save_rejected_records(demandes_traiter_df, request)
            demandes_traiter_df = demandes_traiter_df[
                (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                (demandes_traiter_df['date_debut_retraite'].isna())
            ]

            if demandes_traiter_df.empty:
                messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                return redirect('list_generated')

            # Calculate A
            from dateutil.relativedelta import relativedelta
            demandes_traiter_df['A'] = demandes_traiter_df['date_embauche'].apply(lambda x: relativedelta(datetime.now(), x).years * 12 + relativedelta(datetime.now(), x).months)

            # Calculate S
            def calculate_S(row):
                sit_fam = row['sit_fam'].strip().lower()
                nbr_enf = row['NOMBRE_ENF']

                if sit_fam not in ['célibataire']:
                    if nbr_enf <= 3:
                        return 5 + nbr_enf * 5
                    else:
                        return 20
                else:
                    return 0

            demandes_traiter_df['S'] = demandes_traiter_df.apply(calculate_S, axis=1)

            # Calculate D
            def calculate_D(row, historique_df):
                matricule = row['Matricule']
                matricule_rows = historique_df[historique_df['Matricule'] == matricule]
                total_D = 0

                for index, hist_row in matricule_rows.iterrows():
                    date_de_debut_sejour = hist_row['Date debut sejour']
                    year_difference = datetime.now().year - date_de_debut_sejour.year

                    if year_difference == 1:
                        total_D += 140
                    elif year_difference == 2:
                        total_D += 90
                    elif year_difference == 3:
                        total_D += 50
                    elif year_difference >= 4:
                        total_D += 20

                return total_D

            demandes_traiter_df['D'] = demandes_traiter_df.apply(calculate_D, axis=1, historique_df=historique_df)
            demandes_traiter_df['D'] = demandes_traiter_df['D'].fillna(0)

            # Calculate P
            demandes_traiter_df['P'] = 2 * (demandes_traiter_df['A'] + demandes_traiter_df['S']) - demandes_traiter_df['D']

            # Sorting based on 'P'
            demandes_traiter_df = demandes_traiter_df.sort_values(by=['P', 'A', 'S', 'Date de la demande'], ascending=[False, False, False, False])

            # Save the DataFrame to the database
            save_to_database(demandes_traiter_df,request)
            messages.success(request, "Le traitement des fichiers a été effectué avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue : {str(e)}")

    return redirect('list_generated')  # Replace with the actual view name





def upload_files_and_rank(request):
    if request.method == 'POST':
        form = UploadFilesForm(request.POST, request.FILES)
        if form.is_valid():
            agents_file = request.FILES.get('agents_file')
            historique_file = request.FILES.get('historique_file')
            demandes_file = request.FILES.get('demandes_file')

            # Process the agents file
            if agents_file:
                if not agents_file.name.endswith('.xlsx'):
                    messages.error(request, "Format de fichier invalide pour les agents. Veuillez télécharger un fichier .xlsx.")
                else:
                    try:
                        df = pd.read_excel(agents_file)
                        expected_columns = ['matricule', 'nom_prenom', 'date_naissance', 'sit_fam', 'date_embauche', 'nombre_enf', 'date_debut_retraite']
                        df.columns = df.columns.str.lower()
                        if not all(col in df.columns for col in expected_columns):
                            messages.error(request, "Format de fichier Excel invalide pour les agents. Les colonnes attendues sont manquantes.")
                        else:
                            Agent.objects.all().delete()
                            for index, row in df.iterrows():
                                if pd.isnull(row['matricule']) or pd.isnull(row['nom_prenom']) or pd.isnull(row['date_naissance']):
                                    messages.error(request, f"Données requises manquantes à la ligne {index + 1}.")
                                    return redirect('AgentsAffichage')
                                try:
                                    agent = Agent(
                                        matricule=row['matricule'],
                                        nom_prenom=row['nom_prenom'],
                                        date_naissance=row['date_naissance'] if not pd.isna(row['date_naissance']) else None,
                                        sit_fam=row['sit_fam'],
                                        date_embauche=row['date_embauche'] if not pd.isna(row['date_embauche']) else None,
                                        nombre_enf=row['nombre_enf'],
                                        date_debut_retraite=row.get('date_debut_retraite') if not pd.isna(row.get('date_debut_retraite')) else None
                                    )
                                    agent.save()
                                except Exception as e:
                                    messages.error(request, f"Erreur lors de l'enregistrement des données des agents : {str(e)}")
                                    return redirect('AgentsAffichage')
                            messages.success(request, "Fichier agents téléchargé et traité avec succès !")
                    except Exception as e:
                        messages.error(request, f"Erreur lors de la lecture du fichier Excel des agents : {str(e)}")
                        return redirect('AgentsAffichage')

            # Process the demandes file
            if demandes_file:
                try:
                    df = pd.read_excel(demandes_file)
                    required_columns = ['Ville', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites', 'Statut', 'Nature Periode', 'Site', 'N° Demande']
                    if not all(column in df.columns for column in required_columns):
                        messages.error(request, f'Colonne manquante dans le fichier Excel des demandes.')
                        return redirect('demande_list')
                    Demande.objects.all().delete()
                    for _, row in df.iterrows():
                        Demande.objects.create(
                            ville=row.get('Ville', ''),
                            nom_agent=row.get('Nom agent', ''),
                            prenom_agent=row.get('Prenom agent', ''),
                            matricule=row.get('Matricule', ''),
                            date_demande=row.get('Date de la demande'),
                            date_debut_sejour=row.get('Date debut sejour'),
                            date_fin_sejour=row.get('Date fin sejour'),
                            type_de_vue=row.get('Type de vue', 'Inconnu'),
                            nombre_nuites=row.get('Nombre de nuites', 0),
                            statut=row.get('Statut', ''),
                            nature_periode=row.get('Nature Periode', ''),
                            site=row.get('Site', ''),
                            numero_demande=row.get('N° Demande', '')
                        )
                    messages.success(request, 'Fichier demandes téléchargé et données insérées avec succès.')
                except Exception as e:
                    messages.error(request, f'Erreur lors du traitement du fichier Excel des demandes: {e}')
                    return redirect('demande_list')

            # Process the historique file
            if historique_file:
                if not historique_file.name.endswith('.xlsx'):
                    messages.error(request, "Format de fichier invalide pour l'historique. Veuillez télécharger un fichier .xlsx.")
                else:
                    try:
                        df = pd.read_excel(historique_file)
                        required_columns = ['Site', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites']
                        if not all(column in df.columns for column in required_columns):
                            messages.error(request, "Le fichier Excel de l'historique n'a pas les colonnes requises.")
                            return redirect('historique')
                        Historique.objects.all().delete()
                        rows_with_missing_values = df[df[required_columns].isnull().any(axis=1)]
                        df = df.dropna(subset=required_columns)
                        inserted_rows = 0
                        for _, row in df.iterrows():
                            try:
                                Historique.objects.create(
                                    ville=row['Site'],
                                    nom_agent=row['Nom agent'],
                                    prenom_agent=row['Prenom agent'],
                                    matricule=row['Matricule'],
                                    date_demande=row['Date de la demande'],
                                    date_debut_sejour=row['Date debut sejour'],
                                    date_fin_sejour=row['Date fin sejour'],
                                    type_de_vue=row['Type de vue'],
                                    nombre_nuites=row['Nombre de nuites'],
                                )
                                inserted_rows += 1
                            except Exception as e:
                                messages.error(request, f"Erreur lors de l'insertion de la ligne {row}: {e}")
                        if not rows_with_missing_values.empty:
                            messages.warning(request, "Certaines lignes avaient des valeurs manquantes et n'ont pas été insérées.", extra_tags='missing-values')
                        else:
                            messages.success(request, f"Fichier Excel historique traité avec succès. {inserted_rows} lignes insérées.")
                    except Exception as e:
                        messages.error(request, f"Erreur lors de la lecture du fichier Excel de l'historique : {str(e)}")
                        return redirect('historique')

            # Perform filtering and ranking
            if request.POST.get('filter_and_rank'):
                try:
                    date_debut_sejour = form.cleaned_data['date_debut_sejour']
                    date_fin_sejour = form.cleaned_data['date_fin_sejour']

                    demandes = Demande.objects.filter(
                        date_debut_sejour__gte=date_debut_sejour,
                        date_fin_sejour__lte=date_fin_sejour
                    )

                    historique_queryset = Historique.objects.all().values()
                    agents_queryset = Agent.objects.all().values()
                    demandes_queryset = Demande.objects.all().values()

                    historique_df = pd.DataFrame(list(historique_queryset))
                    agents_df = pd.DataFrame(list(agents_queryset))
                    demandes_df = pd.DataFrame(list(demandes_queryset))

                    historique_df.rename(columns={
                        'ville': 'Ville',
                        'nom_agent': 'Nom agent',
                        'prenom_agent': 'Prenom agent',
                        'matricule': 'Matricule',
                        'date_demande': 'Date de la demande',
                        'date_debut_sejour': 'Date debut sejour',
                        'date_fin_sejour': 'Date fin sejour',
                        'nombre_nuites': 'Nombre de nuites'
                    }, inplace=True)

                    agents_df.rename(columns={
                        'matricule': 'Matricule',
                        'nom_prenom': 'Nom Prenom',
                        'date_naissance': 'Date Naissance',
                        'date_embauche': 'Date Embauche',
                        'nombre_enf': 'Nombre Enfants',
                        'date_debut_retraite': 'Date Debut Retraite'
                    }, inplace=True)

                    demandes_df.rename(columns={
                        'numero_demande': 'Numero Demande',
                        'ville': 'Ville',
                        'nom_agent': 'Nom agent',
                        'prenom_agent': 'Prenom agent',
                        'matricule': 'Matricule',
                        'date_demande': 'Date de la demande',
                        'date_debut_sejour': 'Date debut sejour',
                        'date_fin_sejour': 'Date fin sejour',
                        'type_de_vue': 'Type de vue',
                        'nombre_nuites': 'Nombre de nuites',
                        'statut': 'Statut',
                        'nature_periode': 'Nature Periode',
                        'site': 'Site'
                    }, inplace=True)

                    demandes_df = demandes_df[
                        (demandes_df['Date debut sejour'] >= date_debut_sejour) &
                        (demandes_df['Date fin sejour'] <= date_fin_sejour)
                    ]

                    demandes_traiter_df = demandes_df.copy()
                    demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
                    demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Bloquée']
                    demandes_traiter_df = demandes_traiter_df.groupby('Nom agent')['Nombre de nuites'].sum().reset_index()
                    demandes_traiter_df = demandes_traiter_df.sort_values(by='Nombre de nuites', ascending=False)

                    messages.success(request, "Filtrage et classement effectués avec succès.")
                    # Redirect or render the appropriate template with results
                    return redirect('result_page')  # Replace 'result_page' with your result view name

                except Exception as e:
                    messages.error(request, f"Erreur lors du traitement des données : {str(e)}")
                    return redirect('result_page')  # Replace 'result_page' with your result view name

    else:
        form = UploadFilesForm()

    return render(request, 'upload_files.html', {'form': form})






def upload_rank(request):
    if request.method == 'POST':
        form = UploadFilesFormLibre(request.POST, request.FILES)
        if form.is_valid():
            agents_file = request.FILES.get('agents_file')
            historique_file_first = request.FILES.get('historique_file_first')
            historique_file_second = request.FILES.get('historique_file_second')
            demandes_file = request.FILES.get('demandes_file')
            try:
                # Process agents file
                if agents_file:
                    agents_df = pd.read_excel(agents_file)
                    expected_columns = ['matricule', 'nom_prenom', 'date_naissance', 'sit_fam', 'date_embauche', 'nombre_enf', 'date_debut_retraite']
                    agents_df.columns = agents_df.columns.str.lower()
                    missing_columns = [col for col in expected_columns if col not in agents_df.columns]
                    if missing_columns:
                        messages.error(request, f"Colonnes manquantes dans le fichier agents : {', '.join(missing_columns)}")
                        return redirect('AgentsAffichage')

                    for index, row in agents_df.iterrows():
                        if pd.isnull(row['matricule']) or pd.isnull(row['nom_prenom']) or pd.isnull(row['date_naissance']):
                            messages.error(request, f"Données requises manquantes à la ligne {index + 1}.")
                            return redirect('AgentsAffichage')
                    messages.success(request, "Fichier agents téléchargé et traité avec succès !")

                # Process demandes file
                if demandes_file:
                    demandes_df = pd.read_excel(demandes_file)
                    required_columns = ['Ville', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites', 'Statut', 'Nature Periode', 'Site', 'N° Demande']
                    missing_columns = [col for col in required_columns if col not in demandes_df.columns]
                    if missing_columns:
                        messages.error(request, f'Colonnes manquantes dans le fichier demandes : {", ".join(missing_columns)}')
                        return redirect('demande_list')

                # Process historique files
                if historique_file_first and historique_file_second:
                    if not historique_file_first.name.endswith('.xlsx') or not historique_file_second.name.endswith('.xlsx'):
                        messages.error(request, "Format de fichier invalide pour l'historique. Veuillez télécharger des fichiers .xlsx.")
                        return redirect('historique')

                    historique_df_first = pd.read_excel(historique_file_first)
                    historique_df_second = pd.read_excel(historique_file_second)

                    if set(historique_df_first.columns) != set(historique_df_second.columns):
                        messages.error(request, "Erreur : Les colonnes des deux fichiers historiques ne correspondent pas.")
                        return redirect('historique')

                    try:
                        historique_df = pd.concat([historique_df_first, historique_df_second], ignore_index=True)
                    except ValueError as e:
                        messages.error(request, f"Erreur lors de la concaténation : {e}")
                        return redirect('historique')
                    except Exception as e:
                        messages.error(request, f"Erreur inattendue lors de la concaténation : {e}")
                        return redirect('historique')

                    required_columns = ['Site', 'Nom agent', 'Prenom agent', 'Matricule', 'Date de la demande', 'Date debut sejour', 'Date fin sejour', 'Type de vue', 'Nombre de nuites']
                    missing_columns = [col for col in required_columns if col not in historique_df.columns]
                    if missing_columns:
                        messages.error(request, f'Colonnes manquantes dans le fichier historique : {", ".join(missing_columns)}')
                        return redirect('historique')

                    historique_df = historique_df.dropna(subset=required_columns)
                    if historique_df.empty:
                        messages.warning(request, "Toutes les lignes du fichier historique contenaient des valeurs manquantes et n'ont pas été insérées.")

                # Additional processing after validation
                if form.is_valid():
                    date_debut_sejour = pd.to_datetime(form.cleaned_data.get('date_debut_sejour'))
                    date_fin_sejour = pd.to_datetime(form.cleaned_data.get('date_fin_sejour'))

                    if not date_debut_sejour or not date_fin_sejour:
                        messages.error(request, "Les dates de début et de fin de séjour sont requises.")
                        return redirect('libre_generated')


                # Rename columns for consistency
                historique_df.rename(columns={
                    'ville': 'Ville',
                    'nom_agent': 'Nom agent',
                    'prenom_agent': 'Prenom agent',
                    'matricule': 'Matricule',
                    'date_demande': 'Date de la demande',
                    'date_debut_sejour': 'Date debut sejour',
                    'date_fin_sejour': 'Date fin sejour',
                    'nombre_nuites': 'Nombre de nuites'
                }, inplace=True)

                agents_df.rename(columns={
                    'matricule': 'matricule',
                    'nom_prenom': 'Nom Prenom',
                    'date_naissance': 'Date Naissance',
                    'date_embauche': 'Date Embauche',
                    'nombre_enf': 'Nombre Enfants',
                    'date_debut_retraite': 'date_debut_retraite'
                }, inplace=True)

                demandes_df.rename(columns={
                    'N° Demande': 'Numero Demande',
                    'ville': 'Ville',
                    'nom_agent': 'Nom agent',
                    'prenom_agent': 'Prenom agent',
                    'Matricule': 'Matricule',
                    'date_demande': 'Date de la demande',
                    'date_debut_sejour': 'Date debut sejour',
                    'date_fin_sejour': 'Date fin sejour',
                    'type_de_vue': 'Type de vue',
                    'nombre_nuites': 'Nombre de nuites',
                    'statut': 'Statut',
                    'nature_periode': 'Nature Periode',
                    'site': 'Site'
                }, inplace=True)

                print("*********************matrucule demande")
                
                print(demandes_df['Matricule'])

                demandes_df['Date debut sejour'] = pd.to_datetime(demandes_df['Date debut sejour'], format='%m/%d/%Y', errors='coerce')
                demandes_df['Date fin sejour'] = pd.to_datetime(demandes_df['Date fin sejour'], format='%m/%d/%Y', errors='coerce')


                demandes_df = demandes_df[
                    (demandes_df['Date debut sejour'] >= date_debut_sejour) &
                    (demandes_df['Date fin sejour'] <= date_fin_sejour)
                ]
               
                # Additional filtering
                demandes_traiter_df = demandes_df.copy()
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Site'] == 'Khouribga']
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Nature Periode'] == 'Libre']
                demandes_traiter_df = demandes_traiter_df[demandes_traiter_df['Statut'] == 'En attente de traitement']

                # Convert date columns to datetime
                print(agents_df['Date Embauche'])
                print(agents_df['Date Naissance'])
                print("matricuuule demande avant merge")
                print(demandes_traiter_df['Matricule'])

                demandes_traiter_df['Matricule'] = demandes_traiter_df['Matricule'].astype(str).str.strip()
                agents_df['matricule'] = agents_df['matricule'].astype(str).str.strip()

                print("matricuuule demande traités apres Strip")
                print(demandes_traiter_df['Matricule'])


                # Process and merge data
                demandes_traiter_df = demandes_traiter_df.merge(
                    agents_df[['matricule', 'Date Embauche', 'Nombre Enfants', 'Date Naissance','date_debut_retraite']],
                    left_on='Matricule', right_on='matricule', how='left'
                )
                print("matricuuule demande traités")
                print(demandes_traiter_df['Matricule'])
                print("0000000000000000")
                print(demandes_traiter_df['Date Embauche'])
                print(demandes_traiter_df['Date Naissance'])
                print("11111111111111111")
                demandes_traiter_df['Date Embauche'] = pd.to_datetime(demandes_traiter_df['Date Embauche'], format='%m/%d/%Y', errors='coerce')
                print("2")
                demandes_traiter_df['Date Naissance'] = pd.to_datetime(demandes_traiter_df['Date Naissance'], format='%m/%d/%Y', errors='coerce')
                print("3")
                historique_df['Date debut sejour'] = pd.to_datetime(historique_df['Date debut sejour'], errors='coerce')
                print("4")
                historique_df['Date fin sejour'] = pd.to_datetime(historique_df['Date fin sejour'], errors='coerce')
                demandes_traiter_df['date_debut_retraite'] = pd.to_datetime(demandes_traiter_df['date_debut_retraite'], format='%d/%m/%Y')

                # Filter expired date_debut_retraite
                
                demandes_traiter_df = demandes_traiter_df[
                    (demandes_traiter_df['date_debut_retraite'] > demandes_traiter_df['Date debut sejour']) |
                    (demandes_traiter_df['date_debut_retraite'].isna())
                ]

                if demandes_traiter_df.empty:
                  messages.info(request, "Aucun enregistrement ne correspond aux critères après filtrage.")
                  return redirect('libre_generated')
            
                # Calculate additional fields
                def calculate_anciennete(date_embauche):
                    return relativedelta(datetime.now(), date_embauche).years * 12 + relativedelta(datetime.now(), date_embauche).months

                def calculate_age(date_naissance):
                    return relativedelta(datetime.now(), date_naissance).years
                print("first")
                print(demandes_traiter_df['Matricule'])
                print(demandes_traiter_df['Date Embauche'])
                print(demandes_traiter_df['Date Naissance'])
                print(demandes_traiter_df[['Date Embauche', 'Date Naissance']].isnull().sum())

                demandes_traiter_df['Anciennete'] = demandes_traiter_df['Date Embauche'].apply(calculate_anciennete)
                demandes_traiter_df['Age'] = demandes_traiter_df['Date Naissance'].apply(calculate_age)
                print("second")
                def calculate_sejour_count(matricule, historique_df):
                    return len(historique_df[historique_df['Matricule'] == matricule])

                def calculate_last_sejour(matricule, historique_df):
                    agent_hist = historique_df[historique_df['Matricule'] == matricule]
                    if agent_hist.empty:
                        return pd.NaT
                    return agent_hist['Date debut sejour'].max()

                demandes_traiter_df['Nombre Sejours'] = demandes_traiter_df['Matricule'].apply(lambda x: calculate_sejour_count(x, historique_df))
                demandes_traiter_df['Dernier Sejour'] = demandes_traiter_df['Matricule'].apply(lambda x: calculate_last_sejour(x, historique_df))

                # Handle NaT values in 'Dernier Sejour' column
                demandes_traiter_df['Dernier Sejour'] = demandes_traiter_df['Dernier Sejour'].fillna(pd.Timestamp.min)

                # Sorting agents according to the specified criteria
                def custom_sort(row):
                    if row['Nombre Sejours'] == 0:
                        return (0, row['Date Embauche'], -row['Age'], -row['Nombre Enfants'])
                    else:
                        return (
                            1,  # Priorité aux agents avec des séjours
                            row['Nombre Sejours'],  # Tri par nombre de séjours
                            row['Dernier Sejour'],  # Dernier séjour le plus loin en premier
                            #-row['Anciennete'],  # Ancienneté en mois
                            row['Date Embauche'],
                            -row['Age'],  # Âge
                            -row['Nombre Enfants']  # Nombre d'enfants
                        )

                demandes_traiter_df['Sort Key'] = demandes_traiter_df.apply(custom_sort, axis=1)
                demandes_traiter_df.sort_values(by='Sort Key', inplace=True)
                demandes_traiter_df.drop(columns='Sort Key', inplace=True)
                print("laaaast -1")
                save_rejected_records(demandes_traiter_df, request)
                # Update AgentsLibre with correct Dernier Sejour
                print("laaaast 0")
                AgentsLibre.objects.all().delete()  # Clean up old records
                print("laaaast")
                for index, row in demandes_traiter_df.iterrows():
                    AgentsLibre.objects.create(
                        numero_demande=row['Numero Demande'],
                        ville=row['Ville'],
                        nom_agent=row['Nom agent'],
                        prenom_agent=row['Prenom agent'],
                        matricule=row['Matricule'],
                        date_debut_sejour=row['Date debut sejour'],
                        date_fin_sejour=row['Date fin sejour'],
                        type_de_vue=row['Type de vue'],
                        nombre_enfants=row['Nombre Enfants'],
                        age=row['Age'],
                        anciennete=row['Anciennete'],
                        date_embauche=row['Date Embauche'],
                        nombre_sejour=row['Nombre Sejours'],
                        dernier_sejour=row['Dernier Sejour']
                    )

                messages.success(request, "Les données ont été traitées et sauvegardées avec succès.")
                return redirect('libre_generated')
            except Exception as e:
                messages.error(request, f"Une erreur est survenue: {str(e)}")
                return redirect('libres_files')
    else:
        form = UploadFilesFormLibre()

    return redirect('libre_generated')




def download_pdf_demandes_libre(request):
    demandes = AgentsLibre.objects.all()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading2']
    body_style = styles['BodyText']

    # Title
    title = Paragraph("Liste des Demandes Traitées", title_style)
    elements.append(title)

    # Table header
    data = [
        ["Numéro Demande", "Ville", "Matricule", "Date Début Séjour", "Date Fin Séjour", "Anciennete", 
         "Nombre de séjour","Dernier séjour", "Age", "Nombre Enfants"]
    ]
    
    # Add data
    for demande in demandes:
        data.append([
            demande.numero_demande,
            demande.ville,
            #demande.nom_agent,
            #demande.type_de_vue,
            demande.matricule,
            demande.date_debut_sejour.strftime("%d-%m-%Y"),
            demande.date_fin_sejour.strftime("%d-%m-%Y"),
            demande.anciennete,
            demande.nombre_sejour,
            demande.dernier_sejour,
            demande.age,
            demande.nombre_enfants
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), '#000000'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f5f5f5'),
        ('GRID', (0, 0), (-1, -1), 1, '#000000'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="demandes_traiter.pdf"'
    return response



@require_POST
def delete_demandes_libres(request):
    try:
        AgentsLibre.objects.all().delete()
        messages.success(request, 'Toutes les demandes de periode libre traitées ont été supprimées avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression des demandes traitées: {e}')
    return redirect('libre_generated')  # Redirect to the page displaying the statistics




import pandas as pd
from django.http import HttpResponse
from .models import AgentsLibre

def download_excel_demandes_libre(request):
    # Query the AgentsLibre model
    demandes = AgentsLibre.objects.all()

    # Convert the queryset to a DataFrame
    df = pd.DataFrame(list(demandes.values()))

    # Define columns of interest
    columns_of_interest = [
        'numero_demande', 'ville', 'matricule', 'date_debut_sejour', 
        'date_fin_sejour', 'anciennete', 'nombre_sejour', 
        'dernier_sejour', 'age', 'nombre_enfants'
    ]
    df = df[columns_of_interest]

    # Create a BytesIO buffer to hold the Excel data
    excel_buffer = io.BytesIO()

    # Write the DataFrame to the buffer
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Demandes Traitées')

    # Seek to the beginning of the stream
    excel_buffer.seek(0)

    # Return the Excel file as a downloadable response
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="demandes_traiter.xlsx"'
    
    return response

def download_excel_rejected_demandes(request):

     rejected_demandes = RejectedDemandesRetrait.objects.all()
     df = pd.DataFrame(list(rejected_demandes.values()))
     
     if df.empty:
        # Si le DataFrame est vide, créez un fichier Excel vide ou une feuille de calcul avec des messages appropriés
        wb = Workbook()
        ws = wb.active
        ws.append(['Aucune donnée disponible'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=demandes_rejete.xlsx'
        wb.save(response)
        return response

     data = [
        ["numero_demande", "nom_agent", "prenom_agent", "date_debut_sejour", "date_fin_sejour"]
    ]
     df = df[data]

     

     excel_buffer = io.BytesIO()

    # Write the DataFrame to the buffer
     with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Demandes Traitées')

    # Seek to the beginning of the stream
     excel_buffer.seek(0)

    # Return the Excel file as a downloadable response
     response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
     response['Content-Disposition'] = 'attachment; filename="demandes_rejete.xlsx"'
    
     return response


def download_pdf_libre(request, ville, type_de_vue):
     # Récupérer les données du modèle AgentsLibre selon les critères
    data = AgentsLibre.objects.filter(ville=ville, type_de_vue=type_de_vue).values()
    
    # Convertir les données en DataFrame
    df = pd.DataFrame(data)
    
    # Debug: Afficher les colonnes du DataFrame
    print("DataFrame columns:", df.columns.tolist())
    
    # Colonnes d'intérêt
    columns_of_interest = ['numero_demande', 'matricule',  'nombre_sejour', 
                           'dernier_sejour', 'anciennete', 'age', 'nombre_enfants', 
                            'date_debut_sejour', 'date_fin_sejour']

    # Vérifier si toutes les colonnes existent dans le DataFrame
    missing_cols = [col for col in columns_of_interest if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Colonnes manquantes: {missing_cols}")
    
    # Créer un buffer pour le PDF
    buffer = BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Définir les styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.black,
        spaceAfter=20,
        alignment=1  # Centré
    )
    normal_style = styles['Normal']
    normal_style.fontName = 'Helvetica'
    normal_style.fontSize = 10

    # Titre du document
    title_text = "Rapport des Agents Libres"
    title_paragraph = Paragraph(title_text, title_style)
    elements.append(title_paragraph)
    elements.append(Spacer(1, 12))

    # Convertir le DataFrame en liste de listes pour la Table
    table_data = [columns_of_interest] + df[columns_of_interest].values.tolist()

    # Calculer les largeurs de colonnes basées sur le contenu
    col_widths = [max(len(str(cell)) for cell in column) * 7 for column in zip(*table_data)]
    col_widths = [min(width, 100) for width in col_widths]  # Limiter la largeur maximale à 100

    # Créer la Table
    table = Table(table_data, colWidths=col_widths)
    
    # Appliquer le style de la table
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Couleur de fond pour l'en-tête
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Couleur du texte pour l'en-tête
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Style de police pour l'en-tête
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Taille de la police pour l'en-tête
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),  # Couleur de fond pour les données
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),  # Taille de la police pour les données
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),  # Couleurs alternées pour les lignes
    ])
    table.setStyle(table_style)

    elements.append(table)
    elements.append(Spacer(1, 24))

    # Construire le PDF
    doc.build(elements)
    
    # Récupérer les données PDF du buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Placer le PDF généré dans la réponse HTTP
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{ville}_{type_de_vue}.pdf"'
    
    return response


import pandas as pd
from openpyxl import load_workbook
from .models import AgentsLibre

import pandas as pd
from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO
from .models import AgentsLibre
from openpyxl import load_workbook
from io import BytesIO
from django.http import HttpResponse
from .forms import AffectationForm
from .models import AgentsLibre
from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO

from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO
import itertools

from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO

from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Border, Side, PatternFill

def affecter(request, ville, type_de_vue):
    if request.method == 'POST':
        form = AffectationForm(request.POST, request.FILES)
        if form.is_valid():
            aff_file = request.FILES.get('affectation_file')
            filtred_agents = AgentsLibre.objects.filter(ville=ville, type_de_vue=type_de_vue)
            workbook = load_workbook(aff_file)
            worksheet = workbook.active
            assignments = {}
            
            thick_border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

            # Définir un ensemble de couleurs pour la colorisation
            colors = [
    "FFFF00",  # Jaune
    "00FF00",  # Vert
    "00FFFF",  # Cyan
    "FF00FF",  # Magenta
    "FFA500",  # Orange
    "FFC0CB",  # Rose
    "0000FF",  # Bleu
    "8A2BE2",  # Bleu-violet
    "FFD700",  # Or
    "DC143C",  # Crème
    "4B0082",  # Indigo
    "7FFF00",  # Vert Chartreuse
    "FF4500",  # Rouge orange
]

            color_index = 0
            for agent in filtred_agents:
                assigned = False
                possible_assignments = []

                for row in range(4, worksheet.max_row + 1):
                    chalet = worksheet.cell(row=row, column=2).value
                    if not chalet:
                        continue

                    start_day = agent.date_debut_sejour.day
                    end_day = agent.date_fin_sejour.day
                    period_available = True

                    # Vérification de la disponibilité de la période
                    for day in range(start_day, end_day + 1):
                        if worksheet.cell(row=row, column=day + 2).value:
                            period_available = False
                            break

                    if period_available:
                        # Calcul du gap minimum pour cette période
                        gap_minimum = calculate_gap(worksheet, row, start_day, end_day)
                        possible_assignments.append((row, gap_minimum))

                # Priorisation des affectations selon les critères
                if possible_assignments:
                    possible_assignments.sort(
                        key=lambda x: -x[1] if x[1] is not None else float('inf')  # Prioriser le gap minimum
                    )
                    selected_assignment = possible_assignments[0]
                    selected_row = selected_assignment[0]

                    fill_color = PatternFill(start_color=colors[color_index % len(colors)],
                                             end_color=colors[color_index % len(colors)],
                                             fill_type="solid")
                    color_index += 1

                    # Affecter l'agent au chalet sélectionné
                    for day in range(start_day, end_day + 1):
                        worksheet.cell(row=selected_row, column=day + 2).value = agent.matricule
                        cell = worksheet.cell(row=selected_row, column=day + 2)
                        cell.border = thick_border
                        cell.fill = fill_color
                        
                    
                    assignments[agent.matricule] = worksheet.cell(row=selected_row, column=2).value
                    assigned = True

                if not assigned:
                    print(f"Agent {agent.matricule} non affecté à cause d'un manque de période disponible.")

            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Affectation_{ville}_{type_de_vue}.xlsx"'

            return response
        
        else:
            print("Form errors:", form.errors)
    return redirect('libres_files')

def calculate_gap(worksheet, row, start_day, end_day):
    previous_filled = None
    next_filled = None

    # Trouver la dernière date libre avant start_day
    for day in range(start_day - 1, 2, -1):
        if not worksheet.cell(row=row, column=day + 2).value:
            previous_filled = day + 2
            break

    # Trouver le jour suivant rempli après end_day
    for day in range(end_day + 1, 34):
        if worksheet.cell(row=row, column=day + 2).value:
            next_filled = day + 2
            break

    # Calculer le gap minimum
    if previous_filled and next_filled:
        return min(next_filled - end_day, start_day - previous_filled)
    elif previous_filled:
        return start_day - previous_filled
    elif next_filled:
        return next_filled - end_day
    else:
        return float('inf')  # Période totalement libre, donc gap infini
